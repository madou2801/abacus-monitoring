# -*- coding: utf-8 -*-
"""Génère le classeur de veille « procédures collectives / transfert Mali ».

    python3 genere_excel.py [chemin_sortie.xlsx]

Le classeur produit est vivant : les notes (bleu) et les pondérations (bleu)
sont des cellules de saisie, les scores et les classes sont des formules qui se
recalculent. Aucun score n'est figé en dur.
"""
import csv
import os
import sys
import unicodedata

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from referentiel import CRITERES, SECTEURS  # noqa: E402

ICI = os.path.dirname(os.path.abspath(__file__))

POLICE = "Arial"
BLEU = Font(name=POLICE, size=10, color="0000FF")          # saisie
NOIR = Font(name=POLICE, size=10)                          # formule / texte
GRAS = Font(name=POLICE, size=10, bold=True)
TITRE = Font(name=POLICE, size=14, bold=True, color="1F3864")
SOUS_TITRE = Font(name=POLICE, size=11, bold=True, color="1F3864")
BLANC_GRAS = Font(name=POLICE, size=10, bold=True, color="FFFFFF")

FOND_ENTETE = PatternFill("solid", fgColor="1F3864")
FOND_SAISIE = PatternFill("solid", fgColor="FFF2CC")
FOND_ALERTE = PatternFill("solid", fgColor="FCE4D6")

_fin = Side(style="thin", color="BFBFBF")
CADRE = Border(left=_fin, right=_fin, top=_fin, bottom=_fin)

HAUT = Alignment(vertical="top", wrap_text=True)
CENTRE = Alignment(horizontal="center", vertical="center")


def sans_accent(texte):
    """Clé de rapprochement insensible aux accents et à la casse."""
    decompose = unicodedata.normalize("NFKD", texte)
    return "".join(c for c in decompose if not unicodedata.combining(c)).lower().strip()


NOTES_PAR_SECTEUR = {sans_accent(nom): d["notes"] for nom, d in SECTEURS.items()}
# Libellé canonique (accentué) du référentiel : le CSV peut être saisi sans
# accents, mais la feuille doit porter une seule orthographe, sans quoi les
# formules de synthèse par secteur ne retrouvent plus leurs lignes.
SECTEUR_CANONIQUE = {sans_accent(nom): nom for nom in SECTEURS}


def largeurs(ws, mapping):
    for col, larg in mapping.items():
        ws.column_dimensions[col].width = larg


def entetes(ws, ligne, valeurs):
    for i, val in enumerate(valeurs, start=1):
        c = ws.cell(row=ligne, column=i, value=val)
        c.font = BLANC_GRAS
        c.fill = FOND_ENTETE
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = CADRE
    ws.row_dimensions[ligne].height = 34


# ---------------------------------------------------------------------------
# 1. Lisez-moi
# ---------------------------------------------------------------------------
def feuille_lisezmoi(wb, nb_cas):
    ws = wb.create_sheet("Lisez-moi")
    largeurs(ws, {"A": 26, "B": 108})

    ws["A1"] = "Veille procédures collectives — transférabilité vers le Mali"
    ws["A1"].font = TITRE
    ws["A2"] = "ABACUS RH · généré le"
    ws["A2"].font = NOIR
    ws["B2"] = "=TEXT(TODAY(),\"DD/MM/YYYY\")"
    ws["B2"].font = NOIR

    blocs = [
        (
            "PÉRIMÈTRE DEMANDÉ",
            "Entreprises françaises sous procédure collective (sauvegarde, redressement, "
            "liquidation) dans cinq secteurs : transport & logistique ; fabrication de "
            "matériel électrique ; fabrication de machines-outils ; transformation "
            "plastique ; fabrication de médicaments génériques. Objectif : identifier "
            "celles dont l'activité ou les actifs pourraient être transférés au Mali.",
        ),
        (
            "CE QUE CONTIENT CE FICHIER",
            "Un échantillon de %d dossiers réellement sourcés, chacun accompagné de son "
            "lien de presse ou de base publique, noté sur la grille de transférabilité. "
            "Plus le référentiel complet (37 codes NAF, 7 critères pondérés) et un "
            "modèle de saisie pour l'alimenter." % nb_cas,
        ),
        (
            "CE QU'IL NE CONTIENT PAS",
            "La liste exhaustive. Elle est impossible à produire depuis cet "
            "environnement : le proxy réseau bloque le BODACC, l'INPI, l'API "
            "recherche-entreprises et Gaïa (gaia.abacus-rh.com). Seule la recherche web "
            "généraliste répond, et elle remonte des articles de presse, pas un extrait "
            "de registre. L'ordre de grandeur réel : le seul secteur transport compte "
            "857 défaillances au T1 2026 (+12,9 %) et plus de 1 900 dossiers actifs fin "
            "mars 2026 — l'échantillon ci-contre en couvre une fraction infime.",
        ),
        (
            "COMMENT OBTENIR L'EXHAUSTIF",
            "Le script collecte_bodacc.py, livré à côté de ce fichier, interroge l'API "
            "BODACC Open Data, filtre sur les 37 codes NAF et régénère ce classeur "
            "complet. Il est prêt à tourner : il suffit de l'exécuter depuis le VPS "
            "ABACUS ou toute machine dont le réseau n'est pas restreint, ou d'ajouter "
            "les domaines à l'allowlist de l'environnement Claude Code.",
        ),
        (
            "FIABILITÉ DES DONNÉES",
            "Chaque ligne porte une colonne « Fiabilité source ». Un seul dossier "
            "(Mécanique Gaillonnaise de Précision) dispose d'un SIREN et d'une date de "
            "jugement confirmés ; les autres reposent sur la presse, avec des effectifs "
            "et des dates parfois approximatifs. « n.d. » signifie non disponible et "
            "n'a jamais été comblé par une estimation. Aucune donnée de ce fichier n'est "
            "inventée : vérifiez chaque dossier au registre avant toute offre de reprise.",
        ),
        (
            "EXCLUSIONS ASSUMÉES",
            "Duralex (redressement judiciaire du 01/06/2026, 243 salariés, Orléans) "
            "revient dans toutes les recherches mais relève de la verrerie, hors des "
            "cinq secteurs demandés. Fibre Excellence (pâte à papier) de même. Ni l'un "
            "ni l'autre n'a été retenu.",
        ),
        (
            "LECTURE DES SCORES",
            "Score de 1 à 5, moyenne pondérée des 7 critères de l'onglet « Grille "
            "Mali ». Classe A (>= 4) : à instruire en priorité. Classe B (>= 3) : à "
            "étudier. Classe C (>= 2) : marginal. Classe D (< 2) : non pertinent. Les "
            "notes sont des hypothèses sectorielles, pas des mesures — elles sont en "
            "bleu et destinées à être ajustées dossier par dossier.",
        ),
        (
            "CODE COULEUR",
            "Bleu = cellule de saisie, modifiable. Noir = formule ou texte, ne pas "
            "écraser. Fond orangé = point de vigilance.",
        ),
        (
            "AVERTISSEMENT JURIDIQUE",
            "Une reprise en plan de cession se joue devant le tribunal de commerce, "
            "dans un délai d'offre court (4 mois à compter de l'ouverture, souvent "
            "moins). Le transfert d'actifs hors de France peut se heurter au maintien "
            "de l'emploi, critère légal de sélection des offres, et à un contrôle des "
            "investissements étrangers sur les actifs sensibles — la pharmacie en "
            "particulier. À valider avec un conseil avant tout engagement.",
        ),
    ]

    ligne = 4
    for titre, texte in blocs:
        ws.cell(row=ligne, column=1, value=titre).font = SOUS_TITRE
        c = ws.cell(row=ligne, column=2, value=texte)
        c.font = NOIR
        c.alignment = HAUT
        ws.row_dimensions[ligne].height = max(30, 13 * (len(texte) // 95 + 1))
        if titre in ("CE QU'IL NE CONTIENT PAS", "AVERTISSEMENT JURIDIQUE"):
            ws.cell(row=ligne, column=1).fill = FOND_ALERTE
            c.fill = FOND_ALERTE
        ligne += 2

    ws.freeze_panes = "A4"
    return ws


# ---------------------------------------------------------------------------
# 2. Grille Mali
# ---------------------------------------------------------------------------
def feuille_grille(wb, nb_cas):
    ws = wb.create_sheet("Grille Mali")
    largeurs(ws, {"A": 8, "B": 34, "C": 14, "D": 92, "E": 12, "F": 12,
                  "G": 12, "H": 12, "I": 12})

    ws["A1"] = "Grille de transférabilité vers le Mali"
    ws["A1"].font = TITRE
    ws["A2"] = ("Notes de 1 à 5 : 5 = configuration la plus favorable au transfert. "
                "Les pondérations (bleu) sont modifiables ; leur somme doit rester à 100 %.")
    ws["A2"].font = NOIR

    entetes(ws, 3, ["Code", "Critère", "Pondération", "Définition et justification"])

    for i, (code, nom, poids, definition) in enumerate(CRITERES):
        r = 4 + i
        ws.cell(row=r, column=1, value=code).font = GRAS
        ws.cell(row=r, column=2, value=nom).font = NOIR
        c = ws.cell(row=r, column=3, value=poids)
        c.font = BLEU
        c.fill = FOND_SAISIE
        c.number_format = "0.0%"
        c.alignment = CENTRE
        d = ws.cell(row=r, column=4, value=definition)
        d.font = NOIR
        d.alignment = HAUT
        ws.row_dimensions[r].height = 44
        for col in range(1, 5):
            ws.cell(row=r, column=col).border = CADRE

    total = 4 + len(CRITERES)
    ws.cell(row=total, column=2, value="Total").font = GRAS
    c = ws.cell(row=total, column=3, value="=SUM(C4:C%d)" % (total - 1))
    c.font = GRAS
    c.number_format = "0.0%"
    c.alignment = CENTRE
    ws.cell(row=total, column=4,
            value='=IF(ROUND(C%d,6)=1,"Pondérations cohérentes.",'
                  '"ANOMALIE : la somme des pondérations doit valoir 100 %%.")'
                  % total).font = GRAS

    # Bloc horizontal : indispensable pour SUMPRODUCT contre les notes en ligne.
    l_codes, l_poids = total + 2, total + 3
    ws.cell(row=l_codes, column=2,
            value="Pondérations en ligne (utilisées par les formules de score)").font = GRAS
    for i, (code, _, _, _) in enumerate(CRITERES):
        col = 3 + i
        ws.cell(row=l_codes, column=col, value=code).font = GRAS
        ws.cell(row=l_codes, column=col).alignment = CENTRE
        p = ws.cell(row=l_poids, column=col, value="=C%d" % (4 + i))
        p.font = NOIR
        p.number_format = "0.0%"
        p.alignment = CENTRE
        p.border = CADRE
    ws.cell(row=l_poids, column=2, value="↳ recopie de la colonne C, ne pas saisir ici").font = NOIR

    # Barème
    r = l_poids + 2
    ws.cell(row=r, column=1, value="Barème de classement").font = SOUS_TITRE
    bareme = [
        ("A — Prioritaire", "Score >= 4,0", "Dossier à instruire en priorité."),
        ("B — À étudier", "3,0 <= score < 4,0", "Transfert plausible sous conditions."),
        ("C — Marginal", "2,0 <= score < 3,0", "Seuls certains actifs sont pertinents."),
        ("D — Non pertinent", "Score < 2,0", "Transfert non viable en l'état."),
    ]
    entetes(ws, r + 1, ["Classe", "Seuil", "Lecture"])
    for i, (cl, seuil, lecture) in enumerate(bareme):
        rr = r + 2 + i
        for col, val in ((1, cl), (2, seuil), (3, lecture)):
            cell = ws.cell(row=rr, column=col, value=val)
            cell.font = NOIR
            cell.border = CADRE

    # Synthèse par secteur — moyenne des scores de l'onglet des cas.
    r2 = r + 2 + len(bareme) + 2
    ws.cell(row=r2, column=1, value="Score moyen par secteur (échantillon sourcé)").font = SOUS_TITRE
    entetes(ws, r2 + 1, ["Secteur", "Nb dossiers", "Score moyen", "Classe moyenne"])
    derniere = 1 + nb_cas
    for i, nom in enumerate(SECTEURS):
        rr = r2 + 2 + i
        ws.cell(row=rr, column=1, value=nom).font = NOIR
        n = ws.cell(row=rr, column=2,
                    value="=SUMPRODUCT(--(TRIM('Cas identifiés'!$A$2:$A$%d)=TRIM($A%d)))"
                          % (derniere, rr))
        m = ws.cell(row=rr, column=3,
                    value="=IFERROR(SUMPRODUCT(--(TRIM('Cas identifiés'!$A$2:$A$%d)=TRIM($A%d)),"
                          "'Cas identifiés'!$T$2:$T$%d)/$B%d,\"n.d.\")"
                          % (derniere, rr, derniere, rr))
        cl = ws.cell(row=rr, column=4,
                     value='=IF(NOT(ISNUMBER($C%d)),"n.d.",'
                           'IF($C%d>=4,"A",IF($C%d>=3,"B",IF($C%d>=2,"C","D"))))'
                           % (rr, rr, rr, rr))
        n.alignment = CENTRE
        m.number_format = "0.00"
        m.alignment = CENTRE
        cl.alignment = CENTRE
        for col in range(1, 5):
            ws.cell(row=rr, column=col).font = NOIR
            ws.cell(row=rr, column=col).border = CADRE

    ws.freeze_panes = "A4"
    return ws


# ---------------------------------------------------------------------------
# 3. Cas identifiés
# ---------------------------------------------------------------------------
COLONNES = [
    ("Secteur", 26), ("Raison sociale", 32), ("SIREN", 12), ("Localisation", 30),
    ("Effectif", 10), ("CA (EUR)", 14), ("Procédure", 24), ("Date jugement", 14),
    ("Statut / échéance", 38), ("Actifs industriels", 40), ("Source (URL)", 34),
    ("Fiabilité source", 18),
]
COL_NOTES_DEBUT = len(COLONNES) + 1          # M
COL_SCORE = COL_NOTES_DEBUT + len(CRITERES)  # T
COL_CLASSE = COL_SCORE + 1                   # U
COL_COMMENT = COL_CLASSE + 1                 # V


def _ligne_formules(ws, r, ligne_poids):
    """Écrit le score pondéré et la classe pour la ligne r."""
    d = get_column_letter(COL_NOTES_DEBUT)
    f = get_column_letter(COL_NOTES_DEBUT + len(CRITERES) - 1)
    pd_ = get_column_letter(3)
    pf = get_column_letter(3 + len(CRITERES) - 1)

    s = ws.cell(row=r, column=COL_SCORE,
                value="=IF(COUNT(%s%d:%s%d)<%d,\"\",SUMPRODUCT(%s%d:%s%d,"
                      "'Grille Mali'!$%s$%d:$%s$%d))"
                      % (d, r, f, r, len(CRITERES), d, r, f, r,
                         pd_, ligne_poids, pf, ligne_poids))
    s.font = GRAS
    s.number_format = "0.00"
    s.alignment = CENTRE
    s.border = CADRE

    sc = get_column_letter(COL_SCORE)
    c = ws.cell(row=r, column=COL_CLASSE,
                value='=IF(NOT(ISNUMBER(%s%d)),"",IF(%s%d>=4,"A — Prioritaire",'
                      'IF(%s%d>=3,"B — À étudier",IF(%s%d>=2,"C — Marginal",'
                      '"D — Non pertinent"))))' % (sc, r, sc, r, sc, r, sc, r))
    c.font = GRAS
    c.alignment = CENTRE
    c.border = CADRE


def feuille_cas(wb, cas, ligne_poids):
    ws = wb.create_sheet("Cas identifiés")

    for i, (_, larg) in enumerate(COLONNES, start=1):
        ws.column_dimensions[get_column_letter(i)].width = larg
    for i in range(len(CRITERES)):
        ws.column_dimensions[get_column_letter(COL_NOTES_DEBUT + i)].width = 6
    ws.column_dimensions[get_column_letter(COL_SCORE)].width = 10
    ws.column_dimensions[get_column_letter(COL_CLASSE)].width = 18
    ws.column_dimensions[get_column_letter(COL_COMMENT)].width = 78

    titres = [n for n, _ in COLONNES] + [c[0] for c in CRITERES] + \
             ["Score /5", "Classe", "Commentaire transfert Mali"]
    entetes(ws, 1, titres)

    for i, (code, nom, _, _) in enumerate(CRITERES):
        ws.cell(row=1, column=COL_NOTES_DEBUT + i).comment = Comment(
            "%s — %s\nNote de 1 à 5 (5 = le plus favorable)." % (code, nom), "ABACUS")

    for j, row in enumerate(cas):
        r = 2 + j
        cle = sans_accent(row["secteur"])
        notes = NOTES_PAR_SECTEUR[cle]
        valeurs = [
            SECTEUR_CANONIQUE[cle], row["raison_sociale"], row["siren"], row["localisation"],
            row["effectif"], row["ca_eur"], row["procedure"], row["date_jugement"],
            row["statut"], row["actifs_industriels"], row["source_url"], row["fiabilite"],
        ]
        for k, val in enumerate(valeurs, start=1):
            if k in (5, 6) and val not in ("", "n.d."):
                val = float(val)
            c = ws.cell(row=r, column=k, value=val)
            c.font = NOIR
            c.alignment = HAUT
            c.border = CADRE
            if k == 6 and isinstance(val, float):
                c.number_format = '#,##0 "€"'
            if k == 5 and isinstance(val, float):
                c.number_format = "#,##0"

        for i, note in enumerate(notes):
            c = ws.cell(row=r, column=COL_NOTES_DEBUT + i, value=note)
            c.font = BLEU
            c.fill = FOND_SAISIE
            c.alignment = CENTRE
            c.border = CADRE

        _ligne_formules(ws, r, ligne_poids)

        c = ws.cell(row=r, column=COL_COMMENT, value=row["commentaire_mali"])
        c.font = NOIR
        c.alignment = HAUT
        c.border = CADRE
        ws.row_dimensions[r].height = 62

    note = ws.cell(row=len(cas) + 3, column=1,
                   value="Échantillon non exhaustif, issu de la presse et de bases "
                         "publiques consultées le 30/08/2026 — voir l'onglet Lisez-moi. "
                         "Les notes en bleu sont des hypothèses sectorielles reprises de "
                         "referentiel.py, à ajuster après visite de site.")
    note.font = Font(name=POLICE, size=9, italic=True)
    ws.merge_cells(start_row=len(cas) + 3, start_column=1,
                   end_row=len(cas) + 3, end_column=COL_COMMENT)

    ws.freeze_panes = "C2"
    ws.auto_filter.ref = "A1:%s%d" % (get_column_letter(COL_COMMENT), len(cas) + 1)
    return ws


# ---------------------------------------------------------------------------
# 4. Codes NAF
# ---------------------------------------------------------------------------
def feuille_naf(wb):
    ws = wb.create_sheet("Codes NAF")
    largeurs(ws, {"A": 36, "B": 12, "C": 86})

    ws["A1"] = "Codes NAF retenus pour le filtrage BODACC"
    ws["A1"].font = TITRE
    ws["A2"] = ("Nomenclature NAF rév. 2. La NAF rév. 3 s'applique aux nouvelles "
                "immatriculations depuis 2026 : remapper avant d'exploiter des jugements "
                "postérieurs.")
    ws["A2"].font = NOIR
    ws["A2"].fill = FOND_ALERTE

    entetes(ws, 3, ["Secteur", "Code NAF", "Libellé"])
    r = 4
    for secteur, d in SECTEURS.items():
        debut = r
        for code, libelle in sorted(d["codes"].items()):
            ws.cell(row=r, column=1, value=secteur).font = NOIR
            ws.cell(row=r, column=2, value=code).font = GRAS
            ws.cell(row=r, column=2).alignment = CENTRE
            ws.cell(row=r, column=3, value=libelle).font = NOIR
            for col in range(1, 4):
                ws.cell(row=r, column=col).border = CADRE
                ws.cell(row=r, column=col).alignment = \
                    CENTRE if col == 2 else HAUT
            r += 1
        ws.merge_cells(start_row=debut, start_column=1, end_row=r - 1, end_column=1)
        ws.cell(row=debut, column=1).alignment = Alignment(
            vertical="center", wrap_text=True)

    ws.cell(row=r + 1, column=1, value="Total").font = GRAS
    ws.cell(row=r + 1, column=2, value="=COUNTA(B4:B%d)" % (r - 1)).font = GRAS
    ws.cell(row=r + 1, column=2).alignment = CENTRE
    ws.freeze_panes = "A4"
    return ws


# ---------------------------------------------------------------------------
# 5. Modèle de saisie
# ---------------------------------------------------------------------------
def feuille_modele(wb, ligne_poids):
    ws = wb.create_sheet("Modèle de saisie")

    for i, (_, larg) in enumerate(COLONNES, start=1):
        ws.column_dimensions[get_column_letter(i)].width = larg
    for i in range(len(CRITERES)):
        ws.column_dimensions[get_column_letter(COL_NOTES_DEBUT + i)].width = 6
    ws.column_dimensions[get_column_letter(COL_SCORE)].width = 10
    ws.column_dimensions[get_column_letter(COL_CLASSE)].width = 18
    ws.column_dimensions[get_column_letter(COL_COMMENT)].width = 78

    titres = [n for n, _ in COLONNES] + [c[0] for c in CRITERES] + \
             ["Score /5", "Classe", "Commentaire transfert Mali"]
    entetes(ws, 1, titres)

    exemple = [
        "Transformation plastique", "EXEMPLE — Plastiques du Rhône SAS", "812345678",
        "Vénissieux (69) — TC de Lyon", 84.0, 11200000.0, "Redressement judiciaire",
        "2026-07-15", "Offres à déposer avant le 15/11/2026",
        "12 presses à injecter 80-450 t, 2 extrudeuses, atelier moules",
        "https://www.bodacc.fr/annonce/exemple", "BODACC — jugement publié",
    ]
    for k, val in enumerate(exemple, start=1):
        c = ws.cell(row=2, column=k, value=val)
        c.font = BLEU
        c.fill = FOND_SAISIE
        c.alignment = HAUT
        c.border = CADRE
        if k == 6:
            c.number_format = '#,##0 "€"'
        if k == 5:
            c.number_format = "#,##0"
    for i, note in enumerate(SECTEURS["Transformation plastique"]["notes"]):
        c = ws.cell(row=2, column=COL_NOTES_DEBUT + i, value=note)
        c.font = BLEU
        c.fill = FOND_SAISIE
        c.alignment = CENTRE
        c.border = CADRE
    _ligne_formules(ws, 2, ligne_poids)
    c = ws.cell(row=2, column=COL_COMMENT,
                value="EXEMPLE — ligne fictive de démonstration du format attendu. "
                      "À supprimer avant exploitation.")
    c.font = BLEU
    c.fill = FOND_SAISIE
    c.alignment = HAUT
    c.border = CADRE
    ws.row_dimensions[2].height = 62

    # Lignes vierges prêtes à l'emploi, formules déjà posées.
    for r in range(3, 33):
        for k in range(1, COL_NOTES_DEBUT + len(CRITERES)):
            cell = ws.cell(row=r, column=k)
            cell.font = BLEU
            cell.border = CADRE
            if k >= COL_NOTES_DEBUT:
                cell.fill = FOND_SAISIE
                cell.alignment = CENTRE
        _ligne_formules(ws, r, ligne_poids)
        ws.cell(row=r, column=COL_COMMENT).font = BLEU
        ws.cell(row=r, column=COL_COMMENT).border = CADRE

    lg = ws.cell(row=35, column=1,
                 value="MODE D'EMPLOI — Saisir uniquement les cellules bleues. Les 7 "
                       "colonnes C1 à C7 attendent une note de 1 à 5 (5 = le plus "
                       "favorable au transfert) ; passer la souris sur l'en-tête d'une "
                       "colonne pour en lire la définition. Les colonnes Score et Classe "
                       "sont des formules : ne pas les écraser. Le score reste vide tant "
                       "que les 7 notes ne sont pas renseignées.")
    lg.font = Font(name=POLICE, size=9, italic=True)
    lg.alignment = HAUT
    ws.merge_cells(start_row=35, start_column=1, end_row=35, end_column=COL_COMMENT)
    ws.row_dimensions[35].height = 44

    ws.freeze_panes = "C2"
    return ws


def main():
    sortie = sys.argv[1] if len(sys.argv) > 1 else os.path.join(
        ICI, "procedures_collectives_transfert_mali.xlsx")

    with open(os.path.join(ICI, "cas_identifies.csv"), encoding="utf-8") as fh:
        cas = list(csv.DictReader(fh, delimiter=";"))

    inconnus = {r["secteur"] for r in cas} - set(NOTES_PAR_SECTEUR)
    inconnus = {s for s in inconnus if sans_accent(s) not in NOTES_PAR_SECTEUR}
    if inconnus:
        raise SystemExit("Secteurs absents du référentiel : %s" % ", ".join(inconnus))

    ligne_poids = 4 + len(CRITERES) + 3  # cf. feuille_grille

    wb = Workbook()
    wb.remove(wb.active)
    feuille_lisezmoi(wb, len(cas))
    feuille_cas(wb, cas, ligne_poids)
    feuille_grille(wb, len(cas))
    feuille_naf(wb)
    feuille_modele(wb, ligne_poids)
    wb.save(sortie)
    print("Classeur écrit : %s (%d dossiers)" % (sortie, len(cas)))


if __name__ == "__main__":
    main()
